import pandas as pd
import openpyxl
from openpyxl.styles import Font, colors
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fills import PatternFill
import excel
from datetime import datetime as dt
import re
import os

def load_matching_workbook():
    # Define a regular expression pattern
    # pattern = re.compile(r'data(\(\d+\))?\.xlsx')
    pattern = re.compile(r'data(\s*\(\d+\))?\.xlsx')

    # Get a list of files in the current directory
    file_list = os.listdir()
    print(file_list)
    # Filter files based on the regular expression
    matching_files = [file for file in file_list if pattern.match(file)]

    # Check if there are matching files
    if matching_files:
        # Load the first matching file
        book = openpyxl.load_workbook(matching_files[0], data_only=True)
        print("A matching file found.")
        # Do further processing with the workbook
        return book
    else:
        print("No matching files found.")
        return None

# Get data from a excel file downloaded from RPO dashboard
# book = openpyxl.load_workbook("data.xlsx", data_only=True)
loaded_workbook = load_matching_workbook() # Call the function to load the workbook


# Select a sheet named Export
sheet_data = loaded_workbook["Export"]
# Mapping data
data = excel.read(sheet_data, (2, 1), (sheet_data.max_row, sheet_data.max_column))

# Convert data to pandas data frame
df = pd.DataFrame(data=data, columns=['Servers', 'VBR', 'Country Lien', 'reftech', 'Backup', 'RPO*-24H', 'Last Backup OK', 'Backup Solutions', 'Backup Status', 'Import Date', 'incident', 'url'])

# Convert all data on the column Lask Backup OK to string
df['Last Backup OK'] = df['Last Backup OK'].astype(str)
# Filter and sort data
data_filtered = df.loc[:, ['Servers', 'VBR', 'Country Lien', 'RPO*-24H', 'Last Backup OK', 'Backup Status', 'incident']].sort_values(['Backup Status', 'VBR']).values.tolist()

# Create a sheet name of today date
sheet_name = dt.today().strftime("%Y-%m-%d")
# Get access to sheet report
book = openpyxl.load_workbook("veeam_rpo.xlsx", data_only=True)
sheet_report = book.create_sheet(sheet_name)

# Write and format the header for the report
topcol = ["Servers", "VBR", "Country", "RPO*-24H", "Last Success", "Backup Status", "Open ticket", "Comment, latest status"]
sheet_report.append(topcol)
# Set the color to fill
fill_color = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
font_format = Font(color="000000", bold=True)

# Apply the color to the range A1 to H1
for col in sheet_report.iter_cols(min_col=1, max_col=8, min_row=1, max_row=1):
    for cell in col:
        cell.fill = fill_color
        cell.font = font_format


# Write all filtered data 
excel.write(sheet_report, data_filtered, "A2")

# Saving the workbook creates the file on disk
book.save("veeam_rpo.xlsx")
print("Book saved.")



