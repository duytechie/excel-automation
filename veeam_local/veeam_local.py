import pandas as pd
import openpyxl
from openpyxl.styles import Font, colors
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fills import PatternFill
import excel
from datetime import datetime as dt
import re
import os


def load_matching_workbook():
    # Define a regular expression pattern
    # pattern = re.compile(r'data(\(\d+\))?\.xlsx')
    pattern = re.compile(r'excelreport(\s*\(\d+\))?\.xlsx')

    # Get a list of files in the current directory
    file_list = os.listdir()
    print(file_list)
    # Filter files based on the regular expression
    matching_files = [file for file in file_list if pattern.match(file)]

    # Check if there are matching files
    if matching_files:
        # Load the first matching file
        book = openpyxl.load_workbook(matching_files[0], data_only=True)
        print("A matching file found.")
        # Do further processing with the workbook
        return book
    else:
        print("No matching files found.")
        return None

# Get data from a excel file downloaded from RPO dashboard
# book = openpyxl.load_workbook("data.xlsx", data_only=True)
loaded_workbook = load_matching_workbook() # Call the function to load the workbook

# Select a sheet named Export
sheet_data = loaded_workbook["Report data"]
# Mapping data
data = excel.read(sheet_data, (2, 1), (sheet_data.max_row, sheet_data.max_column))

# Convert data to pandas data frame
df = pd.DataFrame(data=data, columns=['Name', 'Type', 'Platform', 'VBR', 'Status', 'Latest Run', 'Next Run', 'Description'])
# print(df)

# Filter and sort data
data_filtered = df.loc[(df["Next Run"] != "Disabled"), ["Name", "VBR"]]
# Add more columns to complete the sheet
# Define default values
default_values = {
    "Backup Server": None,
    "Status": "Failed",
    "Last Success": None,
    "Oasis Ticket": None,
    "Comment": None
}

# Use assign to create new columns with default values
data_filtered = data_filtered.assign(**default_values)

# Display the modified DataFrame
# print(data_filtered)

# Sort data & convert data to a list
data_to_write = data_filtered.sort_values("VBR").values.tolist()
# print(data_to_write)

# Create a sheet name of today date
sheet_name = dt.today().strftime("%Y-%m-%d")
# Get access to sheet report
book = openpyxl.load_workbook("veeam_local.xlsx", data_only=True)
sheet_report = book.create_sheet(sheet_name)

# Write and format the header for the report
topcol = ["Job Name", "VBR", "Backup Server", "Status", "Last Success", "Open ticket", "Comment"]
sheet_report.append(topcol)
# Set the color to fill
fill_color = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
font_format = Font(color="000000", bold=True)

# Apply the color to the range A1 to H1
for col in sheet_report.iter_cols(min_col=1, max_col=7, min_row=1, max_row=1):
    for cell in col:
        cell.fill = fill_color
        cell.font = font_format


# Write all filtered data 
excel.write(sheet_report, data_to_write, "A2")

# Saving the workbook creates the file on disk
book.save("veeam_local.xlsx")
print("Book saved.")
